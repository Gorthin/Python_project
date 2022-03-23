import openpyxl

# Tworzymy nowy skoroszyt:
nowy_skoroszyt =  openpyxl.Workbook()           # Tworzymy pusty skoroszyt
nowy_arkusz = nowy_skoroszyt.active             # Ustawiamy się na aktywny arkusz
nowy_arkusz.title="mój arkusz 1"                # Zmiana nazwy domyślnej zakładki "Sheet"
nowy_skoroszyt.create_sheet("mój arkusz 2")     # Tworzymy własne zakładki
nowy_skoroszyt.create_sheet("mój arkusz 3")
nowy_arkusz["A1"] = "Dzień dobry!"
nowy_skoroszyt.save("Inne/nowy.xlsx")

skoroszyt  = openpyxl.load_workbook("Inne/test.xlsx") # Otwórz skoroszyt z dysku
arkusze=skoroszyt.get_sheet_names()
print("Lista arkuszy w skoroszycie Excela:", arkusze)  # Lista: ['Rok-2021', 'Arkusz3', 'Wykresy']
# Analizujemy otwarty skoroszyt:
arkusz= skoroszyt["Arkusz3"]  # Ustawiamy się na "Arkusz3"
# lub:
arkusz= skoroszyt[arkusze[1]]                              # "Arkusz3" to drugi element na liście arkuszy

# Usuwamy 3 kolumny, rozpoczynajac od drugiej (w arkuszu pozostanie zatem, pierwsza i piąta kolumna)
arkusz.delete_cols(2, 3)
arkusz.insert_cols(2)    # Dołóż nową kolumnę przed kolumną nr 2

# Operacje na wierszach
arkusz.delete_rows(2, 2) # Usuwamy 2 wiersze, zaczynając od drugiego
arkusz.insert_rows(1, 2) # Dołóż dwa nowe wiersze przed wierszem pierwszym

skoroszyt.save("Inne/test2.xlsx")


