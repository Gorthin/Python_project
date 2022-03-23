import openpyxl

#skoroszyt  = openpyxl.load_workbook("Inne/raport.xlsx") # Otwórz skoroszyt z dysku
skoroszyt  = openpyxl.load_workbook("Inne/raport.xlsx", data_only=True) # Otwórz w trybie danych (bez formuł)

arkusz = skoroszyt["Arkusz1"]                           # Otwórz 'Arkusz1'


#arkusz = skoroszyt.active # Otwiera aktywny arkusz, tj. ten aktualnie "otwarty" czyli używany
pusta= arkusz["B2"]
k0=arkusz['C3']    #  Wartość '230,5'
k1=arkusz['B13']   #  'Data raportu:'
k1_inaczej=arkusz.cell(row=13, column=2)
k2=arkusz['C12']   #  Formuła '=SUM(C3:C11)'
k3=arkusz['C13']   # Data wpisana do komórki
print("Komórka [B2]=", pusta.value)
print(f"Komórka [C3]= '{k0.value}', typ={type(k0.value)}")
print(f"Komórka [B13]='{k1.value}', typ={type(k1.value)}")
print(f"Komórka [B13] adresowana jako row=13, column=2: '{k1_inaczej.value}'")
print(f"Komórka [C12]='{k2.value}', typ={type(k2.value)}")
print(f"Komórka [C13]='{k3.value}', typ={type(k3.value)}")

print("Statystyki dotyczące dokumentu Excela i otwartego skoroszytu:")
print("  Aktywny skoroszyt:", skoroszyt.active)
print(arkusz.calculate_dimension())
print("  Arkusze:", skoroszyt.worksheets)
print("  Obszar danych:", arkusz.dimensions)
print(f"  Wiersze obszaru danych: od {arkusz.min_row}. do {arkusz.max_row}.")
print(f"  Kolumny obszaru danych: od {arkusz.min_column}. do {arkusz.max_column}.")

print("Uzyskiwanie wartości z zakresów komórek:")
print ("  Zakres_komorek B3:F3= ", end = " ")
dane_w   = [arkusz.cell(row=3, column=i).value for i in range(2,7)] # Wartości: od 3., do 6.
print(dane_w)

print ("  Zakres_komorek C11:C14= ", end = " ")
dane_col = [arkusz.cell(row=i,column=3).value for i in range(11,15)]
print(dane_col)

print("Wydruk obszaru danych arkusza z obszaru C2:F3:")
for wiersz in range (2, 4): # 2 wiersze
    for kolumna in range(3, 7): # 4 kolumny
        print("  ", arkusz.cell(row=wiersz, column=kolumna).value, end = " ")
    print()



